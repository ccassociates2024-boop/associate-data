"""
Associate — AI-powered accounting automation for CAs
Converts bank statement PDFs into structured, accounting-ready Excel files.
"""

import io
import re
import pandas as pd
import pdfplumber
import streamlit as st

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Associate · Bank Statement Processor",
    page_icon="🏦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────
# GLOBAL CSS — fintech SaaS dashboard theme
# ─────────────────────────────────────────────
st.markdown("""
<style>
/* ── Reset & base ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: #0f1117; color: #e2e8f0; }
.block-container { padding: 0 2rem 4rem 2rem !important; max-width: 1280px; }

/* ── Scrollbar ── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #1a1f2e; }
::-webkit-scrollbar-thumb { background: #334155; border-radius: 3px; }

/* ── Navbar ── */
.navbar {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1.2rem 0 1.4rem 0;
    border-bottom: 1px solid #1e293b;
    margin-bottom: 2rem;
}
.nav-logo {
    display: flex;
    align-items: center;
    gap: 10px;
}
.nav-logo-icon {
    width: 36px; height: 36px;
    background: linear-gradient(135deg, #6366f1, #8b5cf6);
    border-radius: 10px;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.1rem;
}
.nav-logo-text {
    font-size: 1.35rem;
    font-weight: 800;
    color: #f1f5f9;
    letter-spacing: -0.5px;
}
.nav-logo-text span { color: #818cf8; }
.nav-badge {
    font-size: 0.65rem;
    font-weight: 700;
    background: #1e293b;
    color: #64748b;
    border: 1px solid #334155;
    border-radius: 20px;
    padding: 3px 10px;
    letter-spacing: 0.08em;
    text-transform: uppercase;
}
.nav-tagline {
    font-size: 0.82rem;
    color: #475569;
    font-weight: 500;
}

/* ── Section headers ── */
.section-header {
    display: flex;
    align-items: center;
    gap: 10px;
    margin: 2rem 0 1rem 0;
}
.step-pill {
    background: linear-gradient(135deg, #6366f1, #8b5cf6);
    color: white;
    font-size: 0.7rem;
    font-weight: 700;
    border-radius: 20px;
    padding: 3px 10px;
    letter-spacing: 0.06em;
    text-transform: uppercase;
}
.section-title {
    font-size: 1rem;
    font-weight: 700;
    color: #e2e8f0;
}
.section-sub {
    font-size: 0.8rem;
    color: #64748b;
    margin-top: 0.2rem;
}

/* ── Upload zone ── */
[data-testid="stFileUploadDropzone"] {
    background: #111827 !important;
    border: 2px dashed #334155 !important;
    border-radius: 14px !important;
    padding: 2rem !important;
    transition: border-color 0.2s;
}
[data-testid="stFileUploadDropzone"]:hover {
    border-color: #6366f1 !important;
}

/* ── Metric cards (SaaS style) ── */
div[data-testid="metric-container"] {
    background: #111827;
    border: 1px solid #1e293b;
    border-radius: 14px;
    padding: 1.1rem 1.3rem;
    position: relative;
    overflow: hidden;
}
div[data-testid="metric-container"]::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, #6366f1, #8b5cf6);
}
[data-testid="metric-container"] label {
    color: #64748b !important;
    font-size: 0.75rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.05em !important;
    text-transform: uppercase !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #f1f5f9 !important;
    font-size: 1.6rem !important;
    font-weight: 800 !important;
    letter-spacing: -0.5px !important;
}

/* ── Debit metric — red accent ── */
.metric-debit div[data-testid="metric-container"]::before {
    background: linear-gradient(90deg, #ef4444, #f87171);
}
/* ── Credit metric — green accent ── */
.metric-credit div[data-testid="metric-container"]::before {
    background: linear-gradient(90deg, #10b981, #34d399);
}

/* ── Custom stat card ── */
.stat-card {
    background: #111827;
    border: 1px solid #1e293b;
    border-radius: 14px;
    padding: 1.2rem 1.4rem;
    position: relative;
    overflow: hidden;
}
.stat-card-accent { position: absolute; top:0; left:0; right:0; height:2px; }
.stat-card-label { font-size:0.72rem; font-weight:700; color:#64748b; text-transform:uppercase; letter-spacing:0.06em; margin-bottom:0.4rem; }
.stat-card-value { font-size:1.6rem; font-weight:800; color:#f1f5f9; letter-spacing:-0.5px; }
.stat-card-sub   { font-size:0.75rem; color:#475569; margin-top:0.3rem; }

/* ── Success banner ── */
.success-banner {
    background: linear-gradient(135deg, #064e3b, #065f46);
    border: 1px solid #059669;
    border-radius: 12px;
    padding: 1rem 1.4rem;
    display: flex;
    align-items: center;
    gap: 12px;
    margin: 1rem 0;
}
.success-icon { font-size: 1.4rem; }
.success-text { font-size: 0.9rem; font-weight: 600; color: #6ee7b7; }
.success-sub  { font-size: 0.78rem; color: #34d399; margin-top: 2px; }

/* ── Info box ── */
.info-box {
    background: #0c1428;
    border: 1px solid #1e3a5f;
    border-radius: 12px;
    padding: 1.2rem 1.5rem;
    display: flex;
    align-items: flex-start;
    gap: 14px;
    margin: 0.5rem 0;
}
.info-icon { font-size: 1.6rem; margin-top: 2px; }
.info-title { font-size: 0.9rem; font-weight: 700; color: #93c5fd; margin-bottom: 3px; }
.info-body  { font-size: 0.82rem; color: #64748b; line-height: 1.6; }

/* ── Table ── */
[data-testid="stDataFrame"], [data-testid="stDataEditor"] {
    border-radius: 12px !important;
    overflow: hidden !important;
    border: 1px solid #1e293b !important;
}
/* Keep data editor text visible in dark theme */
.stDataEditor [data-testid="glideDataEditor"] { background: #111827 !important; }
.stDataEditor canvas { background: #111827 !important; }
/* Streamlit dataframe table text */
[data-testid="stDataFrame"] table { color: #e2e8f0 !important; background: #111827 !important; }
[data-testid="stDataFrame"] th { background: #1e293b !important; color: #818cf8 !important; font-weight:700; }
[data-testid="stDataFrame"] td { color: #cbd5e1 !important; border-color: #1e293b !important; }

/* ── Divider ── */
.divider { border: none; border-top: 1px solid #1e293b; margin: 1.8rem 0; }

/* ── Download button ── */
.stDownloadButton > button {
    background: linear-gradient(135deg, #6366f1, #8b5cf6) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 0.65rem 1.8rem !important;
    font-size: 0.9rem !important;
    font-weight: 700 !important;
    letter-spacing: 0.02em !important;
    transition: opacity 0.2s, transform 0.15s !important;
    box-shadow: 0 4px 15px rgba(99,102,241,0.3) !important;
}
.stDownloadButton > button:hover {
    opacity: 0.9 !important;
    transform: translateY(-1px) !important;
}

/* ── Category badge pills ── */
.badge {
    display: inline-block;
    border-radius: 20px;
    padding: 2px 10px;
    font-size: 0.75rem;
    font-weight: 700;
    letter-spacing: 0.03em;
}
.badge-tax      { background:#312e16; color:#fbbf24; }
.badge-expense  { background:#2d1515; color:#f87171; }
.badge-purchase { background:#1e2a4a; color:#93c5fd; }
.badge-income   { background:#132d21; color:#6ee7b7; }
.badge-uncat    { background:#1e293b; color:#94a3b8; }

/* ── Spinner ── */
[data-testid="stSpinner"] p { color: #818cf8 !important; }

/* ── Streamlit misc overrides ── */
.stCaption { color: #475569 !important; }
.stAlert   { border-radius: 10px !important; }
p, li { color: #94a3b8; }
h1,h2,h3 { color: #f1f5f9; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# NAVBAR
# ─────────────────────────────────────────────
st.markdown("""
<div class="navbar">
  <div class="nav-logo">
    <div class="nav-logo-icon">🏦</div>
    <div>
      <div class="nav-logo-text">Associate<span>.</span></div>
    </div>
    <div class="nav-badge">Beta</div>
  </div>
  <div class="nav-tagline">AI-powered accounting automation for CAs</div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SECTION 1 — PDF EXTRACTION (logic)
# ─────────────────────────────────────────────

# All known column name variants across Indian banks
# (HDFC, SBI, ICICI, Axis, Kotak, Yes Bank, PNB, BOB, Canara, federal…)
DATE_KEYWORDS    = ["date", "dt", "txn date", "tran date", "value dt",
                    "value date", "posting date", "trans date", "transaction date"]
DESC_KEYWORDS    = ["description", "narration", "particulars", "particular",
                    "narr", "desc", "details", "remarks", "transaction details",
                    "transaction remarks"]
DEBIT_KEYWORDS   = ["debit", "withdrawal", "withdraw", "debit amt",
                    "debit amount", "withdrawal amt", "amount debited",
                    "withdrawals", "paid out", "debit(inr)", "debit(₹)"]
CREDIT_KEYWORDS  = ["credit", "deposit", "credit amt", "credit amount",
                    "deposit amt", "amount credited", "deposits",
                    "paid in", "credit(inr)", "credit(₹)"]
BALANCE_KEYWORDS = ["balance", "bal", "closing balance", "closing bal",
                    "running balance", "available balance", "ledger balance"]
# Single signed-amount column (negative=Debit, positive=Credit)
AMOUNT_KEYWORDS  = ["amount", "net amount", "net amt", "txn amount",
                    "transaction amount", "amount(inr)", "tran amount",
                    "debit/credit", "net transaction"]
# Dr/Cr type indicator column
DRCP_KEYWORDS    = ["dr/cr", "cr/dr", "type", "txn type", "transaction type",
                    "d/c", "dc", "debit credit", "cr dr"]


def _col_matches(col_text: str, keywords: list[str]) -> bool:
    """Check if a column header matches any of the given keyword patterns."""
    col_clean = col_text.lower().strip()
    return any(kw in col_clean for kw in keywords)


def extract_raw_frames(uploaded_file: io.BytesIO) -> list[pd.DataFrame]:
    """
    Extract raw table frames from PDF using 3 strategies per page:
    1. Default pdfplumber table extraction
    2. Text-based strategy (catches borderless tables)
    3. Raw text line parsing fallback
    """
    frames = []
    strategies = [
        {},  # default
        {"vertical_strategy": "text", "horizontal_strategy": "text",
         "snap_tolerance": 5, "join_tolerance": 5},
        {"vertical_strategy": "lines", "horizontal_strategy": "lines"},
    ]
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            found = False
            for s in strategies:
                try:
                    tables = page.extract_tables(s) if s else page.extract_tables()
                    for t in tables:
                        if t and len(t) > 1:
                            frames.append(pd.DataFrame(t))
                            found = True
                except Exception:
                    continue
                if found:
                    break
            if not found:
                text = page.extract_text() or ""
                parsed = _parse_text_to_rows(text)
                if parsed is not None:
                    frames.append(parsed)
    return [f for f in frames if f is not None and len(f) > 0]


# Keep old name as alias so calling code doesn't break
def extract_tables_from_pdf(uploaded_file: io.BytesIO) -> list[pd.DataFrame]:
    return extract_raw_frames(uploaded_file)


def _parse_text_to_rows(text: str) -> pd.DataFrame | None:
    """Parse raw text into rows using date-line detection."""
    date_re = re.compile(
        r"^(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}"
        r"|\d{1,2}\s+(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+\d{2,4}"
        r"|\d{4}[\/\-]\d{2}[\/\-]\d{2})",
        re.IGNORECASE,
    )
    rows = [re.split(r"\s{2,}", ln.strip())
            for ln in text.splitlines()
            if date_re.match(ln.strip())]
    if not rows:
        return None
    max_c = max(len(r) for r in rows)
    rows  = [r + [""] * (max_c - len(r)) for r in rows]
    return pd.DataFrame(rows, columns=[f"col_{i}" for i in range(max_c)])


def _score_row(row_values: list) -> int:
    """Score a row by how many bank-column keywords it contains."""
    text = " ".join(str(v).lower() for v in row_values if v)
    all_kw = DATE_KEYWORDS + DESC_KEYWORDS + DEBIT_KEYWORDS + CREDIT_KEYWORDS + BALANCE_KEYWORDS
    return sum(1 for kw in all_kw if kw in text)


def _find_header(df: pd.DataFrame) -> int | None:
    """Return index of the best-matching header row (score >= 2), else None."""
    best_idx, best_score = None, 0
    for i, row in df.iterrows():
        s = _score_row(list(row.values))
        if s > best_score:
            best_score, best_idx = s, i
    return best_idx if best_score >= 2 else None


def _assign_col(col_name: str) -> str | None:
    """Map a raw column name to one of 7 standard names, or None to drop."""
    c = str(col_name).lower().strip()
    if _col_matches(c, DATE_KEYWORDS):    return "Date"
    if _col_matches(c, DESC_KEYWORDS):    return "Description"
    if _col_matches(c, DEBIT_KEYWORDS):   return "Debit"
    if _col_matches(c, CREDIT_KEYWORDS):  return "Credit"
    if _col_matches(c, BALANCE_KEYWORDS): return "Balance"
    if _col_matches(c, AMOUNT_KEYWORDS):  return "SignedAmount"  # single ±amount col
    if _col_matches(c, DRCP_KEYWORDS):    return "TxnType"       # Dr/Cr indicator col
    return None


def _extract_col_positions(df: pd.DataFrame, header_idx: int) -> dict[int, str]:
    """
    From the header row at header_idx, return a dict mapping
    column position → standard name.  Only the FIRST match per
    standard name is kept so we never get duplicate columns.
    """
    pos_map: dict[int, str] = {}
    used: set[str] = set()
    for i, v in enumerate(df.iloc[header_idx]):
        flat = re.sub(r"\s+", " ", str(v)).strip() if v else ""
        std  = _assign_col(flat)
        if std and std not in used:
            pos_map[i] = std
            used.add(std)
    return pos_map


ALL_STD_COLS = ["Date", "Description", "Debit", "Credit",
                "Balance", "SignedAmount", "TxnType"]


def _apply_col_positions(df: pd.DataFrame,
                         pos_map: dict[int, str]) -> pd.DataFrame | None:
    """
    Apply a column-position → standard-name map to a data DataFrame.
    Returns a DataFrame with all 7 standard columns (missing ones = "").
    """
    result = {}
    for pos, std in pos_map.items():
        if pos < len(df.columns) and std not in result:
            result[std] = df.iloc[:, pos].reset_index(drop=True)
    if "Date" not in result:
        return None
    for req in ALL_STD_COLS:
        if req not in result:
            result[req] = ""
    out = pd.DataFrame(result)[ALL_STD_COLS]
    # Drop repeated header rows printed on every page
    is_hdr = out["Date"].astype(str).str.lower().str.match(
        r"^(date|dt|txn|tran|value|posting|sl\.|s\.no)"
    )
    out = out[~is_hdr].dropna(how="all").reset_index(drop=True)
    return out if not out.empty else None


def normalize_raw_tables(frames: list[pd.DataFrame]) -> pd.DataFrame | None:
    """
    Strategy:
    1. Scan ALL frames to find the best header and learn the column layout.
    2. Accept layout even if it only has Date + SignedAmount (single-amount format).
    3. Apply that SAME position map to every frame including continuation pages.
    4. Concat and clean.
    """
    # ── Step 1: discover master column layout ────────────────────────────────
    master_pos_map: dict[int, str] | None = None

    for df in frames:
        h = _find_header(df)
        if h is not None:
            candidate = _extract_col_positions(df, h)
            vals = set(candidate.values())
            # Accept if we have Date + any amount indicator
            has_amount = bool(vals & {"Debit", "Credit", "SignedAmount"})
            if candidate and "Date" in vals and has_amount:
                master_pos_map = candidate
                break

    # ── Step 2: positional fallback ──────────────────────────────────────────
    if master_pos_map is None:
        for df in frames:
            if df.empty or len(df.columns) < 3:
                continue
            first = df.iloc[:, 0].astype(str)
            hits  = first.str.match(
                r"^\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}|^\d{4}[\/\-]\d{2}[\/\-]\d{2}"
            ).sum()
            if hits >= 2:
                master_pos_map = _detect_amount_positions(df)
                if master_pos_map:
                    break
        if not master_pos_map:
            return None

    num_expected_cols = max(master_pos_map.keys()) + 1

    # ── Step 3: apply layout to every frame ──────────────────────────────────
    parts = []
    for df in frames:
        try:
            h = _find_header(df)
            data = df.iloc[h + 1:].copy().reset_index(drop=True) if h is not None else df.copy()

            # If column count mismatches by 1-2 (some banks omit Chq col on
            # continuation pages), try shifting the map to align on Date col
            if len(data.columns) != num_expected_cols and len(data.columns) >= 3:
                # Re-detect via positional for this specific frame
                first = data.iloc[:, 0].astype(str)
                hits  = first.str.match(
                    r"^\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4}|^\d{4}[\/\-]\d{2}[\/\-]\d{2}"
                ).sum()
                if hits >= 2:
                    local_map = _detect_amount_positions(data)
                    norm = _apply_col_positions(data, local_map)
                else:
                    norm = None
            else:
                norm = _apply_col_positions(data, master_pos_map)

            if norm is not None and not norm.empty:
                parts.append(norm)
        except Exception:
            continue

    if not parts:
        return None

    combined = pd.concat(parts, ignore_index=True)
    # Force exact column names regardless of what came through
    if list(combined.columns) != ALL_STD_COLS:
        combined = combined.reindex(columns=ALL_STD_COLS, fill_value="")

    # Drop rows where both Date and Description are empty/null
    d_empty = combined["Date"].astype(str).str.strip().isin(["", "None", "nan"])
    n_empty = combined["Description"].astype(str).str.strip().isin(["", "None", "nan"])
    combined = combined[~(d_empty & n_empty)].reset_index(drop=True)

    return combined if not combined.empty else None


def _detect_amount_positions(df: pd.DataFrame) -> dict[int, str]:
    """
    For PDFs with no header, infer column roles by numeric density.
    Also detects signed-amount columns (contain negative values).
    """
    pos_map: dict[int, str] = {0: "Date", 1: "Description"}
    n = len(df.columns)
    if n < 3:
        return pos_map

    def num_score(col_idx: int) -> float:
        vals = df.iloc[:, col_idx].astype(str)
        numeric = vals.str.replace(r"[,₹\s\-\+\(\)]", "", regex=True).str.match(r"^\d+\.?\d*$")
        return numeric.sum() / max(len(vals), 1)

    def has_negatives(col_idx: int) -> bool:
        vals = df.iloc[:, col_idx].astype(str).str.strip()
        return vals.str.startswith("-").any() or vals.str.startswith("(").any()

    numeric_cols = sorted(
        [i for i in range(2, n) if num_score(i) > 0.3],
        key=lambda i: num_score(i), reverse=True
    )

    if not numeric_cols:
        return pos_map

    # Check if any numeric column has mixed +/- signs → single signed amount col
    signed_cols = [i for i in numeric_cols if has_negatives(i)]

    if signed_cols:
        # Single signed-amount format: negative=Debit, positive=Credit
        pos_map[signed_cols[0]] = "SignedAmount"
        # Remaining high-numeric col is likely Balance
        remaining = [i for i in numeric_cols if i != signed_cols[0]]
        if remaining:
            pos_map[remaining[-1]] = "Balance"
    elif len(numeric_cols) >= 3:
        pos_map[numeric_cols[-1]] = "Balance"
        pos_map[numeric_cols[0]]  = "Debit"
        pos_map[numeric_cols[1]]  = "Credit"
    elif len(numeric_cols) == 2:
        pos_map[numeric_cols[0]]  = "Debit"
        pos_map[numeric_cols[1]]  = "Balance"
    elif len(numeric_cols) == 1:
        pos_map[numeric_cols[0]]  = "Debit"

    return pos_map


# ─────────────────────────────────────────────
# SECTION 2 — DATA PROCESSING (logic)
# ─────────────────────────────────────────────

def clean_amount(value) -> float:
    """Strip currency symbols / commas and convert to float."""
    try:
        cleaned = re.sub(r"[^\d.]", "", str(value))
        return float(cleaned) if cleaned else 0.0
    except (ValueError, TypeError):
        return 0.0


def _parse_signed(raw: str) -> tuple[float, str] | None:
    """
    Parse a signed amount string into (magnitude, "Debit"/"Credit").
    Negative or parenthesised → Debit. Positive → Credit.
    Returns None if value is zero or unparseable.
    """
    s = str(raw).strip().replace(",", "").replace("₹", "").replace(" ", "")
    if not s or s in ("nan", "None", "-", "+"):
        return None
    # Parenthesised negative: (1234.56)
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    is_neg = s.startswith("-")
    magnitude = clean_amount(s)
    if magnitude == 0.0:
        return None
    return (magnitude, "Debit" if is_neg else "Credit")


def standardize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert raw columns → Date | Description | Amount | Type.

    Handles THREE formats:
    ① Two-column  — separate Debit / Credit columns (ICICI, Axis, SBI)
    ② Signed-amount — single Amount col, negative=Debit positive=Credit
    ③ Amount + TxnType — single Amount col + Dr/Cr indicator column
    """
    rows = []
    for _, row in df.iterrows():
        date_val  = str(row.get("Date",        "")).strip()
        desc_val  = str(row.get("Description", "")).strip()
        if not date_val or not desc_val:
            continue
        if date_val in ("nan", "None") or desc_val in ("nan", "None"):
            continue

        base = {"Date": date_val, "Description": desc_val}

        signed_raw = str(row.get("SignedAmount", "")).strip()
        txntype_raw = str(row.get("TxnType", "")).strip().lower()
        debit  = clean_amount(row.get("Debit",  ""))
        credit = clean_amount(row.get("Credit", ""))

        # ── Format ③: Amount column + Dr/Cr indicator ────────────────────────
        if signed_raw and signed_raw not in ("", "nan", "None") and txntype_raw:
            magnitude = clean_amount(signed_raw)
            if magnitude > 0:
                is_dr = (txntype_raw.startswith("d") or
                         "dr" in txntype_raw or "debit" in txntype_raw or
                         "withdrawal" in txntype_raw)
                txn_type = "Debit" if is_dr else "Credit"
                rows.append({**base, "Amount": magnitude, "Type": txn_type})
            continue

        # ── Format ②: Single signed amount (negative=Debit, positive=Credit) ─
        if signed_raw and signed_raw not in ("", "nan", "None"):
            parsed = _parse_signed(signed_raw)
            if parsed:
                rows.append({**base, "Amount": parsed[0], "Type": parsed[1]})
            continue

        # ── Format ①: Separate Debit / Credit columns ────────────────────────
        if debit > 0 and credit > 0:
            rows.append({**base, "Amount": debit,  "Type": "Debit"})
            rows.append({**base, "Amount": credit, "Type": "Credit"})
        elif debit > 0:
            rows.append({**base, "Amount": debit,  "Type": "Debit"})
        elif credit > 0:
            rows.append({**base, "Amount": credit, "Type": "Credit"})

    result = pd.DataFrame(rows, columns=["Date", "Description", "Amount", "Type"])
    result = result[result["Description"].str.strip().ne("")]
    return result.reset_index(drop=True)


# ─────────────────────────────────────────────
# SECTION 3 — AI CLASSIFICATION (rule-based)
# ─────────────────────────────────────────────

CATEGORY_RULES: list[tuple[list[str], str]] = [
    (["gst", "igst", "cgst", "sgst", "tax"],                        "Tax"),
    (["salary", "payroll", "remuneration"],                          "Expense"),
    (["amazon", "flipkart", "myntra", "purchase", "shop", "mart",
      "swiggy", "zomato"],                                           "Purchase"),
    (["neft", "imps", "received", "inward", "rtgs received",
      "transfer in"],                                                 "Income"),
]

LEDGER_MAP: dict[str, str] = {
    "Tax":           "GST Ledger",
    "Expense":       "Expense Ledger",
    "Purchase":      "Purchase Ledger",
    "Income":        "Sales Ledger",
    "Uncategorized": "Suspense Ledger",
}

CATEGORY_COLORS: dict[str, str] = {
    "Tax":           "#fbbf24",
    "Expense":       "#f87171",
    "Purchase":      "#93c5fd",
    "Income":        "#6ee7b7",
    "Uncategorized": "#94a3b8",
}


def classify_description(description: str) -> str:
    """Rule-based keyword classifier. Returns category or 'Uncategorized'."""
    desc_lower = description.lower()
    for keywords, category in CATEGORY_RULES:
        if any(kw in desc_lower for kw in keywords):
            return category
    return "Uncategorized"


def suggest_ledger(category: str) -> str:
    """Map a category to its suggested accounting ledger."""
    return LEDGER_MAP.get(category, "Suspense Ledger")


def apply_classification(df: pd.DataFrame) -> pd.DataFrame:
    """Add Category and Ledger columns."""
    df = df.copy()
    df["Category"] = df["Description"].apply(classify_description)
    df["Ledger"]   = df["Category"].apply(suggest_ledger)
    return df


# ─────────────────────────────────────────────
# SECTION 4 — EXCEL EXPORT
# ─────────────────────────────────────────────

def build_excel(df: pd.DataFrame) -> bytes:
    """Build styled Excel from the final DataFrame."""
    output = io.BytesIO()
    export = pd.DataFrame({
        "Date":        df["Date"],
        "Description": df["Description"],
        "Ledger":      df["Ledger"],
        "Category":    df["Category"],
        "Debit":       df.apply(lambda r: r["Amount"] if r["Type"] == "Debit"   else "", axis=1),
        "Credit":      df.apply(lambda r: r["Amount"] if r["Type"] == "Credit"  else "", axis=1),
    })
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name="Bank Statement")
        ws = writer.sheets["Bank Statement"]
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        # ── White/light professional theme ──────────────────────────────────
        header_fill  = PatternFill("solid", fgColor="1E3A5F")   # dark navy header
        header_font  = Font(bold=True, color="FFFFFF", size=11)  # white text
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        alt_fill     = PatternFill("solid", fgColor="EBF3FB")    # light blue alt rows
        white_fill   = PatternFill("solid", fgColor="FFFFFF")    # white base rows

        # Category-colour fills for the Category column (col D = index 4)
        cat_fills = {
            "Tax":           PatternFill("solid", fgColor="FFF8E1"),
            "Expense":       PatternFill("solid", fgColor="FFEBEE"),
            "Purchase":      PatternFill("solid", fgColor="E3F2FD"),
            "Income":        PatternFill("solid", fgColor="E8F5E9"),
            "Uncategorized": PatternFill("solid", fgColor="F5F5F5"),
        }

        thin  = Side(style="thin",   color="BFCFDF")
        thick = Side(style="medium", color="1E3A5F")
        header_border = Border(bottom=thick)
        row_border    = Border(
            top=thin, bottom=thin, left=thin, right=thin
        )

        col_widths = {"A": 13, "B": 46, "C": 24, "D": 18, "E": 14, "F": 14}
        for col_letter, w in col_widths.items():
            ws.column_dimensions[col_letter].width = w
        ws.row_dimensions[1].height = 22

        # Header row
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = header_border

        # Data rows
        cat_col_idx = export.columns.get_loc("Category") + 1  # 1-based
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            is_alt = (row_idx % 2 == 0)
            for cell in row:
                cell.border    = row_border
                cell.alignment = left_align
                cell.fill      = alt_fill if is_alt else white_fill

            # Colour the Category cell individually
            cat_cell = ws.cell(row=row_idx, column=cat_col_idx)
            cat_val  = str(cat_cell.value or "")
            cat_cell.fill = cat_fills.get(cat_val, white_fill)
            cat_cell.font = Font(bold=True, size=10)
            cat_cell.alignment = center_align

            # Right-align numeric Debit / Credit cells
            for col_letter in ("E", "F"):
                ws[f"{col_letter}{row_idx}"].alignment = Alignment(
                    horizontal="right", vertical="center"
                )

        # Freeze top row
        ws.freeze_panes = "A2"

    return output.getvalue()


# ─────────────────────────────────────────────
# SECTION 5 — UI : UPLOAD
# ─────────────────────────────────────────────
st.markdown("""
<div class="section-header">
  <span class="step-pill">Step 01</span>
  <div>
    <div class="section-title">Upload Bank Statement</div>
    <div class="section-sub">Supports text-based PDF bank statements from any Indian bank</div>
  </div>
</div>
""", unsafe_allow_html=True)

col_upload, col_info = st.columns([3, 2], gap="large")

with col_upload:
    uploaded = st.file_uploader(
        label="",
        type=["pdf"],
        help="Upload a text-based PDF bank statement.",
        label_visibility="collapsed",
    )

with col_info:
    st.markdown("""
    <div class="info-box">
      <div class="info-icon">💡</div>
      <div>
        <div class="info-title">How it works</div>
        <div class="info-body">
          1. Upload your bank statement PDF<br>
          2. AI extracts and classifies transactions<br>
          3. Review &amp; edit categories inline<br>
          4. Export clean Excel for your CA software
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

if uploaded is None:
    st.markdown("""
    <div style="text-align:center; padding: 2.5rem 0; color: #334155;">
      <div style="font-size:2.5rem; margin-bottom:0.5rem;">🏦</div>
      <div style="font-size:0.9rem; font-weight:600; color:#475569;">
        Awaiting bank statement upload…
      </div>
      <div style="font-size:0.78rem; color:#334155; margin-top:0.3rem;">
        Works with HDFC, ICICI, SBI, Axis, Kotak &amp; more
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ─────────────────────────────────────────────
# PROCESSING PIPELINE
# ─────────────────────────────────────────────
with st.spinner("🔍  Reading PDF and extracting transaction tables…"):
    try:
        raw_frames = extract_tables_from_pdf(uploaded)
    except Exception as exc:
        st.error(f"❌  Could not open PDF: {exc}")
        st.stop()

if not raw_frames:
    st.error("❌  No tables detected. Ensure this is a text-based (not scanned) PDF.")
    st.stop()

with st.spinner("⚙️  Normalising columns and standardising data…"):
    raw_df = normalize_raw_tables(raw_frames)

if raw_df is None or raw_df.empty:
    st.error("❌  Could not identify a transaction table. The PDF layout may be non-standard.")
    st.stop()

with st.spinner("🤖  Classifying transactions with AI rules…"):
    std_df   = standardize_dataframe(raw_df)
    if std_df.empty:
        st.error("❌  No valid transactions found. Check the PDF content.")
        st.stop()
    final_df = apply_classification(std_df)

# ── Success banner ──────────────────────────
cat_count   = (final_df["Category"] != "Uncategorized").sum()
uncat_count = (final_df["Category"] == "Uncategorized").sum()
total_txns  = len(final_df)

st.markdown(f"""
<div class="success-banner">
  <div class="success-icon">✅</div>
  <div>
    <div class="success-text">Extraction complete — {total_txns} transactions found</div>
    <div class="success-sub">
      {cat_count} auto-classified &nbsp;·&nbsp; {uncat_count} need review
    </div>
  </div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SECTION 6 — SUMMARY METRICS
# ─────────────────────────────────────────────
total_debit  = final_df[final_df["Type"] == "Debit"]["Amount"].sum()
total_credit = final_df[final_df["Type"] == "Credit"]["Amount"].sum()
net_flow     = total_credit - total_debit

st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown("""
<div class="section-header">
  <span class="step-pill">Summary</span>
  <div class="section-title">Financial Overview</div>
</div>
""", unsafe_allow_html=True)

c1, c2, c3, c4, c5 = st.columns(5, gap="small")

with c1:
    st.markdown(f"""
    <div class="stat-card">
      <div class="stat-card-accent" style="background:linear-gradient(90deg,#6366f1,#8b5cf6);"></div>
      <div class="stat-card-label">Transactions</div>
      <div class="stat-card-value">{total_txns}</div>
      <div class="stat-card-sub">Entries extracted</div>
    </div>""", unsafe_allow_html=True)

with c2:
    st.markdown(f"""
    <div class="stat-card">
      <div class="stat-card-accent" style="background:linear-gradient(90deg,#ef4444,#f87171);"></div>
      <div class="stat-card-label">Total Debit</div>
      <div class="stat-card-value" style="color:#f87171;">₹{total_debit:,.0f}</div>
      <div class="stat-card-sub">Money out</div>
    </div>""", unsafe_allow_html=True)

with c3:
    st.markdown(f"""
    <div class="stat-card">
      <div class="stat-card-accent" style="background:linear-gradient(90deg,#10b981,#34d399);"></div>
      <div class="stat-card-label">Total Credit</div>
      <div class="stat-card-value" style="color:#34d399;">₹{total_credit:,.0f}</div>
      <div class="stat-card-sub">Money in</div>
    </div>""", unsafe_allow_html=True)

with c4:
    net_color = "#34d399" if net_flow >= 0 else "#f87171"
    net_label = "Surplus" if net_flow >= 0 else "Deficit"
    st.markdown(f"""
    <div class="stat-card">
      <div class="stat-card-accent" style="background:linear-gradient(90deg,#0ea5e9,#38bdf8);"></div>
      <div class="stat-card-label">Net Flow</div>
      <div class="stat-card-value" style="color:{net_color};">₹{abs(net_flow):,.0f}</div>
      <div class="stat-card-sub">{net_label}</div>
    </div>""", unsafe_allow_html=True)

with c5:
    cat_pct = int(cat_count / total_txns * 100) if total_txns else 0
    st.markdown(f"""
    <div class="stat-card">
      <div class="stat-card-accent" style="background:linear-gradient(90deg,#f59e0b,#fbbf24);"></div>
      <div class="stat-card-label">Categorized</div>
      <div class="stat-card-value" style="color:#fbbf24;">{cat_pct}%</div>
      <div class="stat-card-sub">{uncat_count} need review</div>
    </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SECTION 7 — EDITABLE TRANSACTION TABLE
# ─────────────────────────────────────────────
st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown("""
<div class="section-header">
  <span class="step-pill">Step 02</span>
  <div>
    <div class="section-title">Review &amp; Edit Transactions</div>
    <div class="section-sub">
      Edit the <strong style="color:#818cf8;">Category</strong> column directly —
      uncategorized rows are highlighted for quick review
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Category legend ──
legend_cols = st.columns(6)
legend_items = [
    ("Tax",           "#312e16", "#fbbf24"),
    ("Expense",       "#2d1515", "#f87171"),
    ("Purchase",      "#1e2a4a", "#93c5fd"),
    ("Income",        "#132d21", "#6ee7b7"),
    ("Uncategorized", "#1e293b", "#94a3b8"),
]
for col, (label, bg, fg) in zip(legend_cols, legend_items):
    col.markdown(
        f'<div style="background:{bg};color:{fg};border-radius:20px;padding:4px 12px;'
        f'text-align:center;font-size:0.72rem;font-weight:700;">{label}</div>',
        unsafe_allow_html=True,
    )

st.write("")

category_options = ["Tax", "Expense", "Purchase", "Income", "Uncategorized"]

# Highlight uncategorized rows in the display copy
def highlight_rows(row):
    if row["Category"] == "Uncategorized":
        return ["background-color: #1a1220; color: #94a3b8"] * len(row)
    return [""] * len(row)

edited_df = st.data_editor(
    final_df[["Date", "Description", "Amount", "Type", "Category", "Ledger"]],
    column_config={
        "Date":        st.column_config.TextColumn("Date",         width="small"),
        "Description": st.column_config.TextColumn("Description",  width="large"),
        "Amount":      st.column_config.NumberColumn("Amount (₹)", format="₹%.2f", width="small"),
        "Type":        st.column_config.TextColumn("Type",         width="small"),
        "Category":    st.column_config.SelectboxColumn(
                           "Category",
                           options=category_options,
                           width="medium",
                       ),
        "Ledger":      st.column_config.TextColumn("Suggested Ledger", width="medium"),
    },
    use_container_width=True,
    num_rows="dynamic",
    hide_index=True,
    key="transaction_editor",
)

# Re-derive ledger after any user edits
edited_df["Ledger"] = edited_df["Category"].apply(suggest_ledger)

# Live update metrics after edits
edited_uncat = (edited_df["Category"] == "Uncategorized").sum()
if edited_uncat > 0:
    st.caption(f"⚠️  {edited_uncat} transaction(s) still uncategorized — will export to Suspense Ledger.")
else:
    st.caption("✅  All transactions categorized.")


# ─────────────────────────────────────────────
# SECTION 8 — CATEGORY BREAKDOWN
# ─────────────────────────────────────────────
st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown("""
<div class="section-header">
  <span class="step-pill">Analysis</span>
  <div class="section-title">Category Breakdown</div>
</div>
""", unsafe_allow_html=True)

col_break, col_type = st.columns(2, gap="large")

with col_break:
    breakdown = (
        edited_df.groupby("Category")["Amount"]
        .agg(["sum", "count"])
        .reset_index()
        .rename(columns={"sum": "Total (₹)", "count": "Txns"})
        .sort_values("Total (₹)", ascending=False)
    )
    st.dataframe(
        breakdown,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Total (₹)": st.column_config.NumberColumn(format="₹%.2f"),
        },
    )

with col_type:
    type_summary = (
        edited_df.groupby("Type")["Amount"]
        .agg(["sum", "count"])
        .reset_index()
        .rename(columns={"sum": "Total (₹)", "count": "Txns"})
    )
    st.dataframe(
        type_summary,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Total (₹)": st.column_config.NumberColumn(format="₹%.2f"),
        },
    )


# ─────────────────────────────────────────────
# SECTION 9 — EXPORT
# ─────────────────────────────────────────────
st.markdown('<hr class="divider">', unsafe_allow_html=True)
st.markdown("""
<div class="section-header">
  <span class="step-pill">Step 03</span>
  <div>
    <div class="section-title">Export to Excel</div>
    <div class="section-sub">Download the ledger-ready file for Tally, Zoho Books, or any CA software</div>
  </div>
</div>
""", unsafe_allow_html=True)

col_dl, col_note = st.columns([2, 3], gap="large")

with col_dl:
    excel_bytes = build_excel(edited_df)
    st.download_button(
        label="⬇️  Download Excel File",
        data=excel_bytes,
        file_name="associate_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

with col_note:
    st.markdown("""
    <div class="info-box" style="padding:0.9rem 1.2rem;">
      <div class="info-icon" style="font-size:1.2rem;">📋</div>
      <div>
        <div class="info-title" style="font-size:0.82rem;">Excel columns</div>
        <div class="info-body">
          Date &nbsp;·&nbsp; Description &nbsp;·&nbsp; Ledger &nbsp;·&nbsp;
          Category &nbsp;·&nbsp; Debit &nbsp;·&nbsp; Credit
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

# ── Footer ──
st.markdown("""
<div style="text-align:center; padding: 3rem 0 1rem 0; color:#334155; font-size:0.75rem;">
  Associate · AI-powered accounting automation for CAs &nbsp;|&nbsp;
  Built with Streamlit &amp; pdfplumber
</div>
""", unsafe_allow_html=True)
